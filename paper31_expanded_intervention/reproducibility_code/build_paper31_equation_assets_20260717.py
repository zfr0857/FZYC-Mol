from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
ANALYSIS = ROOT / "scripts" / "analyze_paper31_expanded_intervention_20260717.py"
OUT = ROOT / "output" / "paper31_expanded_intervention_20260717" / "equation_assets"


EQUATIONS = [
    (1, "Selected candidate", "ĵ_u(K)=arg max_(j∈C_K) V_(u,j)", "selection_units/_selection_record", "selected_candidate"),
    (2, "Observed finite-audit best", "j_best_u(K)=arg max_(j∈C_K) A_(u,j)", "selection_units/_selection_record", "oracle_candidate"),
    (3, "Selection loss", "L_u(K)=A_(u,j_best_u(K))-A_(u,ĵ_u(K))", "_selection_record", "same_unit_selection_gap"),
    (4, "Range-normalized selection loss", "L̃_u(K)=L_u(K)/(max_j A_(u,j)-min_j A_(u,j)+ε)", "_selection_record", "range_normalized_selection_loss"),
    (5, "Chance-adjusted Hit@q", "CAHit_(q,u)=(I(r_u≤q)-q/K)/(1-q/K), q=3", "_selection_record", "chance_adjusted_hit3"),
    (6, "Mean reciprocal rank unit", "MRR_u=1/r_u", "_selection_record", "reciprocal_rank"),
    (7, "Random-order MRR expectation", "E_0(MRR)=H_K/K", "_selection_record", "random_order_mrr_expectation"),
    (8, "Normalized MRR", "MRR̃_u=(MRR_u-H_K/K)/(1-H_K/K)", "_selection_record", "chance_adjusted_mrr"),
    (9, "Leave-one-seed-out reference", "j_ref_(-s)(K)=arg max_j [(1/|U_(-s)|) ∑_(u∈U_(-s)) A_(u,j)]", "add_cross_fitted_gap", "cross_reference_candidate"),
    (10, "Cross-fitted seed-level gap", "L_(CF,s)(K)=(1/F) ∑_(f=1)^F [A_((s,f),j_ref_(-s)(K))-A_((s,f),ĵ_((s,f))(K))]", "add_cross_fitted_gap/bootstrap_summary", "cross_fitted_selection_gap"),
    (11, "Endpoint expansion effect", "Δ_(CF,e)=(1/S) ∑_(s=1)^S [L_(CF,e,s)(32)-L_(CF,e,s)(4)]", "expansion_effects", "delta_CF_endpoint"),
    (12, "Raw audit matrix", "X_(raw,u,j)=A_(u,j)", "effective_diversity", "transformation=raw"),
    (13, "Row-centred matrix", "X_(ctr,u,j)=A_(u,j)-(1/K) ∑_l A_(u,l)", "effective_diversity", "transformation=row_centred"),
    (14, "Fixed-reference matrix", "X_(ref,u,j)=A_(u,j)-A_(u,j_0)", "effective_diversity", "transformation=shared_morgan_linear_relative/fixed_morgan_rf_relative"),
    (15, "Within-unit-rank matrix", "X_(rank,u,j)=rank_j(A_(u,j))", "effective_diversity", "transformation=within_unit_rank"),
    (16, "Ledoit-Wolf covariance", "Σ̂_LW=(1-α)S+αT", "_effective_rank: sklearn.covariance.LedoitWolf", "ledoit_wolf_shrinkage_alpha"),
    (17, "Eigenvalue proportions", "p_i=λ_i/(∑_l λ_l)", "_effective_rank", "eigenvalue_probability_json"),
    (18, "Entropy effective rank", "r_ent=exp[-∑_i p_i ln(p_i)]", "_effective_rank", "entropy_rank"),
    (19, "Participation-ratio rank", "r_PR=(∑_i λ_i)^2/(∑_i λ_i^2)=1/(∑_i p_i^2)", "_effective_rank", "participation_rank"),
    (20, "Observed audit-best opportunity", "G_(best,u,p,K)=max_(j∈C_(p,K)) A_(u,j)-A_(u,a)", "_selection_record", "oracle_opportunity_gain"),
    (21, "Validation-selected gain", "G_(sel,u,p,K)=A_(u,ĵ_(u,p,K))-A_(u,a)", "_selection_record", "selected_model_gain"),
    (22, "Paired homogeneous-normalized selected gain", "G̃_(sel,e,p,K)=1/(SF) ∑_(s,f) G_(sel,(s,f),p,K)/G_(best,(s,f),hom,K)", "add_paired_homogeneous_normalization/bootstrap_summary", "homogeneous_normalized_selected_gain"),
    (23, "Relative entropy rank", "d_(rel,e,p,K)=r_ent(X_(e,p,K))/K", "effective_diversity", "relative_entropy_rank"),
    (24, "Paired normalized cross-fitted gap", "L̃_(CF,e,p,K)=1/(SF) ∑_(s,f) [A_((s,f),j_ref_(-s))-A_((s,f),ĵ_((s,f)))]/G_(best,(s,f),hom,K)", "add_paired_homogeneous_normalization/bootstrap_summary", "homogeneous_normalized_cross_fitted_gap"),
    (25, "Downstream selected-gain efficiency", "η_(down,u,p,K)=G_(sel,u,p,K)/T_(down,u,p,K)", "_selection_record", "downstream_efficiency_per_second"),
    (26, "Selection entropy", "H_sel=-∑_j q_j ln(q_j)", "selection_stability", "candidate_selection_entropy"),
    (27, "Normalized selection entropy", "H_(sel,norm)=H_sel/ln(K)", "selection_stability", "candidate_selection_entropy_normalized"),
]


NOTATION = [
    ("u=(s,f)", "outer audit unit indexed by repeat seed s and outer fold f"),
    ("e", "molecular-property endpoint"),
    ("p", "candidate-pool composition"),
    ("K", "eligible candidate count"),
    ("C_(p,K)", "locked K-candidate prefix for pool p"),
    ("V_(u,j)", "mean inner-validation utility of candidate j in outer unit u"),
    ("A_(u,j)", "outer-audit utility of candidate j in outer unit u"),
    ("a", "predefined anchor; primary analysis uses shared Morgan linear"),
    ("j_0", "reference candidate fixed before outer outcomes are inspected"),
    ("r_u", "rank of the outer-audit best candidate in the inner-validation ordering"),
    ("q_j", "proportion of repeated audit units selecting candidate j"),
    ("S", "number of repeat seeds; S=5"),
    ("F", "number of outer folds per seed; F=3"),
    ("H_K", "Kth harmonic number"),
    ("epsilon", "numerical stabilizer; epsilon=1e-12"),
    ("T_down", "measured downstream inner-plus-outer fit/predict wall time; pretrained encoder cost excluded"),
    ("S (covariance)", "sample covariance matrix in Eq. 16; distinct from the seed-count symbol by context"),
    ("T (shrinkage)", "Ledoit-Wolf scaled-identity shrinkage target"),
    ("alpha", "Ledoit-Wolf shrinkage coefficient estimated analytically by scikit-learn"),
    ("lambda_i", "non-negative eigenvalue of the shrinkage correlation matrix after clipping numerical negatives to zero"),
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    if [number for number, *_ in EQUATIONS] != list(range(1, 28)):
        raise AssertionError("Equations must be numbered consecutively from 1 through 27")
    analysis_text = ANALYSIS.read_text(encoding="utf-8")
    missing = sorted({
        function for _, _, _, function, _ in EQUATIONS
        if function.split("/")[0].split(":")[0] not in analysis_text and not function.startswith("paired K")
    })
    if missing:
        raise AssertionError(f"Unmapped implementation labels: {missing}")

    with (OUT / "Paper31_numbered_equations_linear.txt").open("w", encoding="utf-8") as handle:
        for number, title, equation, _, _ in EQUATIONS:
            handle.write(f"{number}. {title}\n{equation}\n\n")
    with (OUT / "Paper31_equation_code_mapping.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["equation_number", "title", "linear_equation", "implementation", "output_field", "analysis_script_sha256"])
        for number, title, equation, implementation, field in EQUATIONS:
            writer.writerow([number, title, equation, implementation, field, sha256(ANALYSIS)])
    with (OUT / "Paper31_equation_notation.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["symbol", "definition"])
        writer.writerows(NOTATION)
    mapping_frame = pd.DataFrame([
        {
            "equation_number": number, "title": title, "linear_equation": equation,
            "implementation": implementation, "output_field": field,
            "analysis_script_sha256": sha256(ANALYSIS),
        }
        for number, title, equation, implementation, field in EQUATIONS
    ])
    notation_frame = pd.DataFrame(NOTATION, columns=["symbol", "definition"])
    workbook = OUT / "Paper31_equation_notation_and_code_mapping.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        notation_frame.to_excel(writer, sheet_name="Notation", index=False)
        mapping_frame.to_excel(writer, sheet_name="Equation-code mapping", index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F5597")
            for column in range(1, sheet.max_column + 1):
                values = [sheet.cell(row, column).value for row in range(1, sheet.max_row + 1)]
                width = min(max(len(str(value)) for value in values if value is not None) + 2, 48)
                sheet.column_dimensions[get_column_letter(column)].width = max(12, width)
    audit = {
        "status": "complete", "equation_count": len(EQUATIONS), "notation_count": len(NOTATION),
        "epsilon": 1e-12, "equations_22_24": "paired outer-unit normalization before fold-to-seed-to-endpoint aggregation",
        "analysis_script": str(ANALYSIS), "analysis_script_sha256": sha256(ANALYSIS),
    }
    (OUT / "Paper31_equation_asset_audit.json").write_text(json.dumps(audit, indent=2), encoding="utf-8")
    print(json.dumps(audit, indent=2))


if __name__ == "__main__":
    main()
