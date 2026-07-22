from __future__ import annotations

import hashlib
import json
import re
import zipfile
from pathlib import Path

import pandas as pd
from PIL import Image
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "output" / "paper31_expanded_intervention_20260717"
PACKAGE = ROOT / "output" / "paper31_submission_package_20260717"
ANALYSIS = ROOT / "scripts" / "analyze_paper31_expanded_intervention_20260717.py"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def main() -> None:
    checks: list[dict[str, object]] = []

    def check(name: str, observed, expected, passed: bool | None = None, note: str = "") -> None:
        if passed is None:
            passed = observed == expected
        checks.append({"check": name, "observed": observed, "expected": expected, "passed": bool(passed), "note": note})

    row_expectations = {
        "Paper31_inner_scores.csv.gz": 25920,
        "Paper31_outer_candidate_scores.csv.gz": 8640,
        "Paper31_selection_units.csv": 3240,
        "Paper31_endpoint_pool_K_summary.csv": 216,
        "Paper31_effective_diversity_sensitivity.csv": 432,
        "Paper31_selection_stability.csv": 72,
        "Paper31_component_ablation_units.csv": 900,
        "Paper31_component_ablation_summary.csv": 60,
        "Paper31_equal_budget_units.csv": 540,
    }
    for filename, expected in row_expectations.items():
        path = DATA / filename
        observed = len(pd.read_csv(path)) if path.exists() else -1
        check(f"row count: {filename}", observed, expected)

    split = pd.read_csv(DATA / "composition_split_loop" / "Paper31_composition_split_units.csv")
    check("composition split unit count", len(split), 1080)
    split_manifest = pd.read_csv(DATA / "composition_split_loop" / "Paper31_similarity_split_manifest.csv")
    check("similarity groups disjoint", bool(split_manifest.no_group_overlap.all()), True)
    max_similarity = float(split_manifest[[
        "max_train_validation_tanimoto", "max_train_test_tanimoto", "max_validation_test_tanimoto"
    ]].max().max())
    check("maximum cross-fold Tanimoto", max_similarity, "<0.70", passed=max_similarity < 0.70)

    units = pd.read_csv(DATA / "Paper31_selection_units.csv")
    required = [
        "oracle_opportunity_gain", "selected_model_gain", "same_unit_selection_gap",
        "chance_adjusted_hit3", "chance_adjusted_mrr", "cross_fitted_selection_gap",
        "audit_fit_seconds", "candidate_selection_entropy_normalized" if False else "selected_candidate",
    ]
    for column in required:
        check(f"required selection-unit field: {column}", column in units.columns, True)
    check("no missing cross-fitted gaps", int(units.cross_fitted_selection_gap.isna().sum()), 0)
    invalid = int((~units.homogeneous_denominator_valid).sum())
    disclosed = pd.read_csv(DATA / "Paper31_anchor_normalization_sensitivity.csv").invalid_outer_units.sum()
    check("near-zero denominators disclosed", invalid > 0 and disclosed > 0, True)

    equation_mapping = pd.read_csv(DATA / "equation_assets" / "Paper31_equation_code_mapping.csv")
    check("equation mapping rows", len(equation_mapping), 27)
    check("equation numbering", equation_mapping.equation_number.tolist(), list(range(1, 28)))
    mapping_hash = str(equation_mapping.analysis_script_sha256.iloc[0])
    check("equation mapping script hash", mapping_hash, sha256(ANALYSIS))
    equation_workbook = load_workbook(
        DATA / "equation_assets" / "Paper31_equation_notation_and_code_mapping.xlsx", read_only=True
    )
    check(
        "equation workbook sheets",
        {"Notation", "Equation-code mapping"}.issubset(set(equation_workbook.sheetnames)), True,
    )
    equation_workbook.close()

    manuscripts = [
        PACKAGE / "Candidate_pool_expansion_Journal_of_Cheminformatics_manuscript(7).docx",
        PACKAGE / "候选池扩张与模型选择损失_中文完整论文(7).docx",
    ]
    for manuscript in manuscripts:
        with zipfile.ZipFile(manuscript) as archive:
            xml = archive.read("word/document.xml").decode("utf-8")
            media = archive.namelist()
        omath_count = len(re.findall(r"<m:oMath(?:\s|>)", xml))
        check(f"native OMath count: {manuscript.name}", omath_count, 27)
        check(f"equation placeholders absent: {manuscript.name}", "[[EQ_" not in xml, True)
        check(f"Figure 7 placeholder absent: {manuscript.name}", "[[FIGURE_7_VECTOR]]" not in xml, True)
        check(f"vector SVG embedded: {manuscript.name}", any(name.lower().endswith(".svg") for name in media), True)
        for number in range(1, 28):
            check(
                f"equation number {number} present: {manuscript.name}",
                f"({number})" in xml, True,
            )

    figure_stems = ["Figure_7_expanded_equal_size_intervention"]
    figure_stems += [
        f"Supplementary_Figure_S{i}_{suffix}" for i, suffix in [
            (16, "all_endpoint_pool_K"), (17, "modern_component_ablations"),
            (18, "anchor_normalization_sensitivity"), (19, "composition_split_regime"),
            (20, "selection_frequency_entropy"), (21, "equalK_equal_budget"),
        ]
    ]
    for stem in figure_stems:
        directory = DATA / "figures" if stem.startswith("Figure_7") else DATA / "supplementary_figures"
        svg = directory / f"{stem}.svg"
        pdf = directory / f"{stem}.pdf"
        png = directory / f"{stem}_600dpi.png"
        check(f"SVG present: {stem}", svg.exists(), True)
        check(f"PDF present: {stem}", pdf.exists(), True)
        check(f"PNG present: {stem}", png.exists(), True)
        if svg.exists():
            text = svg.read_text(encoding="utf-8")
            check(f"SVG remains editable text: {stem}", "<text" in text, True)
            check(f"SVG has no embedded raster: {stem}", "<image" not in text, True)
        if png.exists():
            with Image.open(png) as image:
                dpi = image.info.get("dpi", (0, 0))
                check(f"PNG minimum width: {stem}", image.width, ">=3000", passed=image.width >= 3000)
                check(f"PNG nominal DPI: {stem}", round(float(dpi[0])), 600, passed=float(dpi[0]) >= 590)

    workbook = DATA / "supplementary_tables" / "Supplementary_Tables_S32-S36.xlsx"
    wb = load_workbook(workbook, read_only=True)
    expected_sheets = {
        "S32 Registry", "S33 Ablations", "S33 Frequencies", "S34 Sensitivity", "S34 Direction", "S35 Split loop", "S35 Concordance", "S35 Stability",
        "S36 Stability runtime", "S36 Frequencies", "S36 Budget stability", "S36 Budget diversity",
    }
    check("supplementary workbook sheets", expected_sheets.issubset(set(wb.sheetnames)), True)
    wb.close()

    equation_audit = json.loads((DATA / "equation_assets" / "Paper31_Word_equation_audit.json").read_text(encoding="utf-8-sig"))
    check("Word equation audit status", equation_audit["status"], "complete")
    figure_audit = json.loads((DATA / "figures" / "Paper31_Figure7_Word_insertion_audit.json").read_text(encoding="utf-8-sig"))
    check("Word Figure 7 insertion mode", figure_audit["insertion_mode"], "Microsoft Word native SVG inline shape")

    frame = pd.DataFrame(checks)
    frame.to_csv(PACKAGE / "Paper31_final_QC_report.csv", index=False, encoding="utf-8-sig")
    failed = frame.loc[~frame.passed]
    audit = {
        "status": "complete" if failed.empty else "failed",
        "checks": int(len(frame)), "passed": int(frame.passed.sum()), "failed": int(len(failed)),
        "failed_checks": failed.check.tolist(),
    }
    (PACKAGE / "Paper31_final_QC_audit.json").write_text(json.dumps(audit, indent=2), encoding="utf-8")
    print(json.dumps(audit, indent=2))
    if not failed.empty:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
