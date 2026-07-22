from __future__ import annotations

import copy
import csv
import os
import tempfile
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\fzyc\output\paper33_final_minor_revision_20260718")
DOCS = [
    (ROOT / "Candidate_pool_expansion_Journal_of_Cheminformatics_FINAL_CLEAN.docx", "en"),
    (ROOT / "候选池扩张与模型选择损失_中文终稿.docx", "zh"),
]
SOURCE_MAPPING = Path(r"D:\fzyc\output\paper31_submission_package_20260717\equation_assets\Paper31_equation_code_mapping.csv")
GROUPS = [[1, 2], [3, 4], [5, 6], [7, 8], [9], [10, 11], [12, 13], [14, 15], [16, 17], [18, 19], [20, 21], [22, 23], [24], [26, 27]]

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
M = "http://schemas.openxmlformats.org/officeDocument/2006/math"
NS = {"w": W, "m": M}
qn = etree.QName


def child(parent, namespace: str, tag: str, **attributes: str):
    node = etree.SubElement(parent, qn(namespace, tag))
    for key, value in attributes.items():
        node.set(qn(namespace, key), value)
    return node


def reference_paragraph(text: str) -> etree._Element:
    paragraph = etree.Element(qn(W, "p"))
    ppr = child(paragraph, W, "pPr")
    child(ppr, W, "spacing", before="0", after="80")
    run = child(paragraph, W, "r")
    rpr = child(run, W, "rPr")
    child(rpr, W, "rFonts", ascii="Times New Roman", hAnsi="Times New Roman", eastAsia="宋体")
    child(rpr, W, "sz", val="20")
    text_node = child(run, W, "t")
    text_node.text = text
    return paragraph


def number_equations(xml: bytes, language: str) -> bytes:
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.fromstring(xml, parser)
    body = tree.find(".//w:body", NS)
    replacements = {
        "en": {
            "Composition, normalization and downstream cost": "Composition and normalization",
            "The label hom denotes the homogeneous Morgan pool, and T_down is measured inner-plus-outer downstream fit/predict wall time.": "The label hom denotes the homogeneous Morgan pool; complete downstream-efficiency and Pareto definitions are provided in Supplementary Methods.",
        },
        "zh": {
            "候选池组成、归一化与下游成本": "候选池组成与归一化",
            "hom 表示同质 Morgan 候选池，T_down 为内外层下游拟合与预测的实测墙钟时间。": "hom表示同质Morgan候选池；完整下游效率与Pareto定义见补充方法。",
        },
    }[language]
    for paragraph in body.findall("w:p", NS):
        texts = paragraph.xpath(".//w:t", namespaces=NS)
        combined = "".join(node.text or "" for node in texts)
        if combined in replacements and texts:
            texts[0].text = replacements[combined]
            for node in texts[1:]:
                node.text = ""
    equations = [
        paragraph
        for paragraph in body.findall("w:p", NS)
        if paragraph.xpath(".//m:oMath", namespaces=NS)
    ]
    if len(equations) != 14:
        raise RuntimeError(f"Expected 14 displayed equations, found {len(equations)}")

    marker = "Equations (1)–(14)" if language == "en" else "公式（1）–（14）"
    full_text = "".join(tree.xpath(".//w:t/text()", namespaces=NS))
    if marker not in full_text:
        sentence = (
            "Equations (1)–(14) define the candidate decisions, selection losses, chance-adjusted ranking, cross-fitted contrasts, matrix transformations, effective-rank summaries, normalized gains and selection stability used below."
            if language == "en"
            else "公式（1）–（14）依次定义下文使用的候选决策、选择损失、机会校正排序、交叉拟合对比、矩阵变换、有效秩、归一化增益与选择稳定性。"
        )
        equations[0].addprevious(reference_paragraph(sentence))

    text_width = 9298
    for number, paragraph in enumerate(equations, start=1):
        math = copy.deepcopy(paragraph.xpath(".//m:oMath", namespaces=NS)[0])
        # Display equation 13 retains the paired cross-fitted gap only.
        # The downstream-efficiency expression is reported in Supplementary Methods.
        if number == 13:
            children = list(math)
            separator_index = next(
                (index for index, node in enumerate(children)
                 if ";" in "".join(node.xpath(".//m:t/text()", namespaces=NS))),
                None,
            )
            if separator_index is not None:
                for node in children[separator_index:]:
                    math.remove(node)
        ppr = paragraph.find("w:pPr", NS)
        if ppr is None:
            ppr = etree.Element(qn(W, "pPr"))
        else:
            ppr = copy.deepcopy(ppr)
        for tag in ("tabs", "jc", "spacing"):
            existing = ppr.find(f"w:{tag}", NS)
            if existing is not None:
                ppr.remove(existing)
        tabs = child(ppr, W, "tabs")
        child(tabs, W, "tab", val="center", pos=str(text_width // 2))
        child(tabs, W, "tab", val="right", pos=str(text_width))
        child(ppr, W, "jc", val="left")
        child(ppr, W, "spacing", before="100", after="100")
        if ppr.find("w:keepLines", NS) is None:
            child(ppr, W, "keepLines")

        for node in list(paragraph):
            paragraph.remove(node)
        paragraph.append(ppr)
        tab1 = child(paragraph, W, "r")
        child(tab1, W, "tab")
        paragraph.append(math)
        tab2 = child(paragraph, W, "r")
        child(tab2, W, "tab")
        run = child(paragraph, W, "r")
        rpr = child(run, W, "rPr")
        child(rpr, W, "rFonts", ascii="Times New Roman", hAnsi="Times New Roman")
        child(rpr, W, "sz", val="20")
        number_text = child(run, W, "t")
        number_text.text = f"({number})"

    return etree.tostring(tree, xml_declaration=True, encoding="UTF-8", standalone="yes")


def update_docx(path: Path, language: str) -> None:
    fd, name = tempfile.mkstemp(suffix=".docx", dir=path.parent)
    os.close(fd)
    temporary = Path(name)
    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "word/document.xml":
                    data = number_equations(data, language)
                target.writestr(item, data)
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def mapping_rows() -> list[dict[str, str]]:
    with SOURCE_MAPPING.open(encoding="utf-8-sig", newline="") as handle:
        source = {int(row["equation_number"]): row for row in csv.DictReader(handle)}
    rows = []
    for display, group in enumerate(GROUPS, start=1):
        items = [source[number] for number in group]
        equations = " ; ".join(item["linear_equation"].replace("q_j", "π_j") for item in items)
        rows.append({
            "display_equation": str(display),
            "source_equations": ", ".join(str(number) for number in group),
            "titles": " / ".join(item["title"] for item in items),
            "linear_equation": equations,
            "implementation": " ; ".join(dict.fromkeys(item["implementation"] for item in items)),
            "output_field": " ; ".join(dict.fromkeys(item["output_field"] for item in items)),
        })
    return rows


def write_mapping(rows: list[dict[str, str]]) -> None:
    csv_path = ROOT / "Equation_to_code_mapping.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equation mapping"
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row[header] for header in headers])
    header_fill = PatternFill("solid", fgColor="D9EAF2")
    for cell in sheet[1]:
        cell.font = Font(name="Times New Roman", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [18, 20, 42, 68, 46, 46]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(ROOT / "Equation_to_code_mapping.xlsx")


def main() -> None:
    for path, language in DOCS:
        update_docx(path, language)
    rows = mapping_rows()
    write_mapping(rows)
    print(f"numbered_equations=14 mapping_rows={len(rows)}")


if __name__ == "__main__":
    main()
