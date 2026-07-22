from __future__ import annotations

import hashlib
import json
import math
import os
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("D:/fzyc")
NEW = Path(os.environ.get("FZYC_ANALYSIS_OUT", ROOT / "output" / "paper22_major_revision_20260713"))
OUT = Path(os.environ.get("FZYC_MINOR_OUT", ROOT / "output" / "paper23_minor_revision_20260713"))
DISPLAY = {
    "bace": "BACE", "bbbp": "BBBP", "clintox": "ClinTox", "esol": "ESOL",
    "freesolv": "FreeSolv", "lipo": "Lipophilicity", "tdc_caco2_wang": "Caco2",
    "tdc_hia_hou": "HIA", "tdc_pgp_broccatelli": "P-gp",
}


def sha(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def q25(x: pd.Series) -> float:
    return float(x.quantile(0.25))


def q75(x: pd.Series) -> float:
    return float(x.quantile(0.75))


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="D9E7F2")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 400) + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(max(map(len, values), default=8) + 2, 10), 42)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)


def add_frame(wb: Workbook, name: str, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet(name[:31])
    ws.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        ws.append([None if pd.isna(v) else v for v in row])
    style_sheet(ws)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    ranking_path = NEW / "chance_adjusted_ranking_units.csv"
    effective_path = NEW / "effective_rank_bootstrap_5000_summary.csv"
    cross_path = NEW / "cross_fitted_complete_intervals.csv"
    matched_path = NEW / "matched_k3_220_subset_summary.csv"
    support_path = NEW / "chemical_support_selection_audit.csv"

    ranking = pd.read_csv(ranking_path)
    expected_seeds = {11, 23, 37, 53, 71}
    expected_folds = {1, 2, 3}
    expected_k = {4, 8, 16, 32}
    assert ranking.task.nunique() == 9
    assert set(ranking.split_seed) == expected_seeds
    assert set(ranking.outer_fold) == expected_folds
    assert set(ranking.candidate_count) == expected_k
    assert len(ranking) == 9 * 5 * 3 * 4
    assert (ranking.groupby(["task", "split_seed", "candidate_count"]).size() == 3).all()

    ranking["mrr_recomputed"] = 1.0 / ranking.oracle_validation_rank
    ranking["mrr_random_expectation"] = ranking.candidate_count.map(
        lambda k: sum(1.0 / i for i in range(1, int(k) + 1)) / int(k)
    )
    ranking["normalized_mrr_gain_recomputed"] = (
        ranking.mrr_recomputed - ranking.mrr_random_expectation
    ) / (1.0 - ranking.mrr_random_expectation)
    max_mrr_error = float((ranking.mrr - ranking.mrr_recomputed).abs().max())
    max_norm_error = float((ranking.normalized_mrr_gain - ranking.normalized_mrr_gain_recomputed).abs().max())
    assert max_mrr_error < 1e-12 and max_norm_error < 1e-12

    metrics = ["chance_adjusted_hit", "normalized_mrr_gain", "mrr", "ndcg", "spearman", "kendall", "rank_percentile"]
    seed = ranking.groupby(
        ["task", "task_type", "candidate_count", "split_seed"], as_index=False
    )[metrics].mean()
    endpoint = seed.groupby(["task", "task_type", "candidate_count"], as_index=False)[metrics].mean()
    main = endpoint.groupby("candidate_count")[metrics].agg(["median", q25, q75, "mean"]).reset_index()
    main.columns = [
        "candidate_count" if a == "candidate_count" else f"{a}_{b}"
        for a, b in main.columns.to_flat_index()
    ]
    sensitivity = endpoint.groupby(["task_type", "candidate_count"])[metrics].agg(["median", q25, q75, "mean"]).reset_index()
    sensitivity.columns = [
        a if not b else f"{a}_{b}" for a, b in sensitivity.columns.to_flat_index()
    ]

    effective = pd.read_csv(effective_path)
    effective_k32 = effective[
        effective.candidate_count.eq(32) & effective.reference_label.eq("candidate_1")
    ].copy()
    effective_verify = effective_k32.groupby("transformation").agg(
        n_endpoints=("task", "nunique"),
        entropy_rank_median=("ledoit_wolf_entropy_rank", "median"),
        entropy_rank_q25=("ledoit_wolf_entropy_rank", q25),
        entropy_rank_q75=("ledoit_wolf_entropy_rank", q75),
        entropy_rank_min=("ledoit_wolf_entropy_rank", "min"),
        entropy_rank_max=("ledoit_wolf_entropy_rank", "max"),
        participation_rank_median=("ledoit_wolf_participation_rank", "median"),
        median_correlation=("ledoit_wolf_median_correlation", "median"),
    ).reset_index()
    assert set(effective_verify.transformation) == {
        "raw", "row_centred", "fixed_reference_relative", "within_unit_rank"
    }

    cross = pd.read_csv(cross_path).copy()
    cross["interval_excludes_zero"] = (
        cross.split_seed_bootstrap95_low_cross_fitted.gt(0)
        | cross.split_seed_bootstrap95_high_cross_fitted.lt(0)
    )
    cross["interpretation"] = np.where(
        ~cross.interval_excludes_zero,
        np.where(cross.cross_fitted_effect.ge(0), "Uncertain, point estimate positive", "Uncertain, point estimate negative"),
        np.where(cross.cross_fitted_effect.gt(0), "Greater loss at K = 32", "Lower loss at K = 32"),
    )
    assert int(cross.cross_fitted_effect.gt(0).sum()) == 6
    assert int(cross.cross_fitted_effect.lt(0).sum()) == 3
    assert set(cross.loc[cross.interval_excludes_zero, "task"]) == {"clintox", "tdc_hia_hou", "esol", "freesolv", "lipo"}

    matched = pd.read_csv(matched_path)
    matched_endpoint = matched.groupby(["task", "task_type"]).agg(
        endpoint_median_gain=("selected_model_gain_median", "median"),
        subset_q25=("selected_model_gain_median", q25),
        subset_q75=("selected_model_gain_median", q75),
        subset_min=("selected_model_gain_median", "min"),
        subset_max=("selected_model_gain_median", "max"),
        positive_subset_proportion=("selected_model_gain_mean", lambda x: float((x > 0).mean())),
        n_subsets=("subset_id", "nunique"),
    ).reset_index()
    assert (matched_endpoint.n_subsets == 220).all()
    assert int(matched_endpoint.endpoint_median_gain.gt(0).sum()) == 8

    support = pd.read_csv(support_path)
    support_summary = support.groupby(["task_type", "tanimoto_bin"]).agg(
        selected_performance_median=("selected_performance", "median"),
        selected_performance_q25=("selected_performance", q25),
        selected_performance_q75=("selected_performance", q75),
        n_fold_units=("selected_performance", "count"),
    ).reset_index()

    hashes = {p.name: sha(p) for p in [ranking_path, effective_path, cross_path, matched_path, support_path]}
    master_rows: list[dict] = []

    def add(metric, level, k, endpoint_name, task_type, estimate, low, high, interval, src, locations, figures, notes=""):
        master_rows.append({
            "metric": metric, "aggregation_level": level, "K": k, "endpoint": endpoint_name,
            "task": task_type, "estimate": estimate, "lower_CI": low, "upper_CI": high,
            "interval_type": interval, "source_file": src.name, "source_hash": hashes[src.name],
            "manuscript_locations": locations, "figure_locations": figures, "notes": notes,
        })

    for r in main.itertuples(index=False):
        for metric in metrics:
            add(
                metric, "fold mean within seed; seed mean within endpoint; median across 9 endpoints",
                int(r.candidate_count), "all endpoints", "task-stratified sensitivity reported separately",
                getattr(r, f"{metric}_median"), getattr(r, f"{metric}_q25"), getattr(r, f"{metric}_q75"),
                "endpoint IQR", ranking_path,
                "Abstract Results; Methods 2.7; Results 3.2; Discussion 4.3; Conclusions",
                "Figure 3A-B", "Random MRR expectation is H_K/K; no pooled outer-unit main estimate.",
            )
    for r in effective_verify.itertuples(index=False):
        add(
            f"Ledoit-Wolf entropy rank: {r.transformation}", "median across 9 endpoints", 32,
            "all endpoints", "classification and regression endpoints retained separately in source",
            r.entropy_rank_median, r.entropy_rank_q25, r.entropy_rank_q75, "endpoint IQR", effective_path,
            "Abstract Results; Results 3.1; Discussion 4.1-4.2; Conclusions", "Figure 2A",
            "Matrix-dependent estimate; not an intrinsic number of independent predictors.",
        )
    for r in cross.itertuples(index=False):
        add(
            "cross-fitted K=32 minus K=4 selection loss", "mean of 3 folds within each of 5 split seeds",
            32, DISPLAY[r.task], r.task_type, r.cross_fitted_effect,
            r.split_seed_bootstrap95_low_cross_fitted, r.split_seed_bootstrap95_high_cross_fitted,
            "split-seed clustered bootstrap 95% CI", cross_path,
            "Abstract Results; Results 3.4; Table 3; Discussion 4.4; Conclusions", "Figure 3C; Figure 4C",
            r.interpretation,
        )
    add("positive cross-fitted direction count", "endpoint count", 32, "all endpoints", "task-stratified", 6, np.nan, np.nan, "not applicable", cross_path, "Abstract Results; Results 3.4; Conclusions", "Figure 3C", "Six positive and three negative endpoint estimates.")
    add("negative cross-fitted direction count", "endpoint count", 32, "all endpoints", "task-stratified", 3, np.nan, np.nan, "not applicable", cross_path, "Abstract Results; Results 3.4; Conclusions", "Figure 3C")
    for r in matched_endpoint.itertuples(index=False):
        add(
            "matched K=3 selected-model gain versus Morgan K=3", "median across all 220 registered subsets",
            3, DISPLAY[r.task], r.task_type, r.endpoint_median_gain, r.subset_q25, r.subset_q75,
            "subset-distribution IQR; not a confidence interval", matched_path,
            "Abstract Results; Results 3.6; Discussion 4.5", "Figure 5A-B",
            "Overlapping subsets are sensitivity contrasts, not independent experiments.",
        )
    add("positive matched-size endpoint-median count", "endpoint count", 3, "all endpoints", "task-stratified", 8, np.nan, np.nan, "not applicable", matched_path, "Abstract Results; Results 3.6", "Figure 5A", "Eight positive and one negative endpoint median.")
    for r in support_summary.itertuples(index=False):
        add(
            "selected performance by chemical support", "median across fold-level audit units", np.nan,
            r.tanimoto_bin, r.task_type, r.selected_performance_median,
            r.selected_performance_q25, r.selected_performance_q75, "fold-unit IQR", support_path,
            "Results 3.8; Discussion 4.6", "Figure 6B", "Classification ROC-AUC and regression RMSE are not pooled.",
        )
    master = pd.DataFrame(master_rows)

    files = {
        "ranking_metric_seed_summary.csv": seed,
        "ranking_metric_endpoint_summary.csv": endpoint,
        "ranking_metric_main_summary.csv": main,
        "ranking_metric_task_stratum_sensitivity.csv": sensitivity,
        "effective_rank_verification.csv": effective_verify,
        "cross_fitted_result_verification.csv": cross,
        "matched_k3_220_subset_verification.csv": matched_endpoint,
        "chemical_support_verification.csv": support_summary,
        "master_result_consistency.csv": master,
    }
    for name, frame in files.items():
        frame.to_csv(OUT / name, index=False)

    audit = {
        "status": "complete",
        "ranking_source": str(ranking_path),
        "ranking_source_sha256": hashes[ranking_path.name],
        "ranking_rows": len(ranking),
        "aggregation": "mean 3 outer folds within split seed; mean 5 split seeds within endpoint; median and IQR across 9 endpoints",
        "random_expected_mrr": "H_K/K",
        "max_absolute_mrr_recomputation_error": max_mrr_error,
        "max_absolute_normalized_mrr_recomputation_error": max_norm_error,
        "legacy_pooled_outer_unit_summary_retired": True,
        "cross_fitted_positive_endpoints": int(cross.cross_fitted_effect.gt(0).sum()),
        "cross_fitted_negative_endpoints": int(cross.cross_fitted_effect.lt(0).sum()),
        "intervals_excluding_zero": [DISPLAY[x] for x in cross.loc[cross.interval_excludes_zero, "task"]],
        "matched_k3_positive_endpoint_medians": int(matched_endpoint.endpoint_median_gain.gt(0).sum()),
        "effective_rank_statement": "Matrix transformation changes the estimand; no single true effective candidate count is asserted.",
        "source_hashes": hashes,
    }
    (OUT / "minor_revision_verification_audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    wb.remove(wb.active)
    add_frame(wb, "Master result table", master)
    add_frame(wb, "Ranking main", main)
    add_frame(wb, "Ranking endpoints", endpoint)
    add_frame(wb, "Ranking seeds", seed)
    add_frame(wb, "Ranking task strata", sensitivity)
    add_frame(wb, "Effective rank", effective_verify)
    add_frame(wb, "Cross-fitted", cross)
    add_frame(wb, "Matched K3", matched_endpoint)
    add_frame(wb, "Chemical support", support_summary)
    provenance = pd.DataFrame([{"source_file": k, "sha256": v} for k, v in hashes.items()])
    add_frame(wb, "Source hashes", provenance)
    wb.save(OUT / "Minor_revision_master_results_and_verification.xlsx")
    print(json.dumps({
        "output": str(OUT), "master_rows": len(master),
        "mrr_K4_median": float(main.loc[main.candidate_count.eq(4), "normalized_mrr_gain_median"].iloc[0]),
        "mrr_K32_median": float(main.loc[main.candidate_count.eq(32), "normalized_mrr_gain_median"].iloc[0]),
        "cross_positive": 6, "cross_negative": 3, "matched_positive": 8,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
