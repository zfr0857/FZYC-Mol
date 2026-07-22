from __future__ import annotations

from pathlib import Path

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.lines import Line2D
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "output" / "paper31_expanded_intervention_20260717"
OUT = DATA / "figures"
SUPP = DATA / "supplementary_figures"
SOURCE = DATA / "figure_source_data"
KS = [4, 8, 16, 32]
TASKS = ["clintox", "bace", "bbbp", "esol", "lipo", "tdc_caco2_wang"]
TASK_LABEL = {
    "clintox": "ClinTox", "bace": "BACE", "bbbp": "BBBP",
    "esol": "ESOL", "lipo": "Lipophilicity", "tdc_caco2_wang": "Caco2",
}
POOLS = ["Homogeneous Morgan", "Classical multiview", "Modern-augmented"]
POOL_SHORT = {
    "Homogeneous Morgan": "Homogeneous",
    "Classical multiview": "Multiview",
    "Modern-augmented": "Modern",
}
COLORS = {
    "Homogeneous Morgan": "#4E79A7",
    "Classical multiview": "#2A9D8F",
    "Modern-augmented": "#E76F51",
}


def style() -> None:
    mpl.rcParams.update({
        "font.family": "Arial",
        "font.size": 8,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "xtick.labelsize": 7,
        "ytick.labelsize": 7,
        "legend.fontsize": 7,
        "axes.linewidth": 0.7,
        "svg.fonttype": "none",
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "savefig.facecolor": "white",
    })


def save(fig: plt.Figure, directory: Path, stem: str) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    fig.savefig(directory / f"{stem}.svg", bbox_inches="tight")
    fig.savefig(directory / f"{stem}.pdf", bbox_inches="tight")
    fig.savefig(directory / f"{stem}_600dpi.png", dpi=600, bbox_inches="tight")
    plt.close(fig)


def panel_label(ax: plt.Axes, label: str) -> None:
    ax.text(-0.08, 1.04, label, transform=ax.transAxes, fontsize=11, fontweight="bold", va="bottom")


def clean(ax: plt.Axes, grid_axis: str = "y") -> None:
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis=grid_axis, color="#D9DEE5", linewidth=0.55, alpha=0.8)
    ax.set_axisbelow(True)


def vector_heatmap(ax: plt.Axes, values: np.ndarray, **kwargs):
    """Draw cells as vector paths so SVG/PDF contains no embedded raster."""
    rows, columns = values.shape
    mesh = ax.pcolormesh(
        np.arange(columns + 1), np.arange(rows + 1), values,
        shading="flat", antialiased=False, **kwargs,
    )
    ax.set_xlim(0, columns)
    ax.set_ylim(rows, 0)
    return mesh


def load() -> dict[str, pd.DataFrame]:
    return {
        "units": pd.read_csv(DATA / "Paper31_selection_units.csv"),
        "summary": pd.read_csv(DATA / "Paper31_endpoint_pool_K_summary.csv"),
        "diversity": pd.read_csv(DATA / "Paper31_effective_diversity_sensitivity.csv"),
        "stability": pd.read_csv(DATA / "Paper31_selection_stability.csv"),
        "ablation": pd.read_csv(DATA / "Paper31_component_ablation_summary.csv"),
        "ablation_diversity": pd.read_csv(DATA / "Paper31_component_ablation_diversity.csv"),
        "normalization": pd.read_csv(DATA / "Paper31_anchor_normalization_sensitivity.csv"),
        "budget_units": pd.read_csv(DATA / "Paper31_equal_budget_units.csv"),
        "budget": pd.read_csv(DATA / "Paper31_equal_budget_summary.csv"),
        "budget_stability": pd.read_csv(DATA / "Paper31_equal_budget_selection_stability.csv"),
        "budget_diversity": pd.read_csv(DATA / "Paper31_equal_budget_effective_diversity.csv"),
    }


def export_figure7_source(data: dict[str, pd.DataFrame]) -> None:
    SOURCE.mkdir(parents=True, exist_ok=True)
    summary = data["summary"]
    primary = summary.loc[
        summary.design.eq("equal_K") & summary.anchor_scheme.eq("shared_morgan_linear")
    ].copy()
    panel_a = primary.loc[primary.candidate_count.eq(32), [
        "pool", "task", "task_type", "candidate_count",
        "homogeneous_normalized_oracle_opportunity_mean",
        "homogeneous_normalized_oracle_opportunity_low",
        "homogeneous_normalized_oracle_opportunity_high",
        "homogeneous_normalized_selected_gain_mean",
        "homogeneous_normalized_selected_gain_low",
        "homogeneous_normalized_selected_gain_high",
    ]]
    panel_b = primary.groupby(["pool", "task_type", "candidate_count"], as_index=False).agg(
        normalized_selected_gain=("homogeneous_normalized_selected_gain_mean", "mean"),
        normalized_cross_fitted_gap=("homogeneous_normalized_cross_fitted_gap_mean", "mean"),
        endpoint_count=("task", "nunique"),
    )
    panel_c = primary[["pool", "task", "task_type", "candidate_count", "chance_adjusted_hit3_mean"]].merge(
        data["stability"].loc[data["stability"].candidate_count.eq(32), [
            "pool", "task", "candidate_count", "candidate_selection_entropy_normalized"
        ]],
        on=["pool", "task", "candidate_count"], how="left",
    )
    equal_k = data["units"].loc[
        data["units"].design.eq("equal_K") & data["units"].anchor_scheme.eq("shared_morgan_linear")
    ].groupby(["pool", "candidate_count"], as_index=False).agg(
        audit_time_seconds=("audit_fit_seconds", "mean"),
        normalized_selected_gain=("homogeneous_normalized_selected_gain", "mean"),
        endpoint_count=("task", "nunique"),
    )
    equal_k["design"] = "equal_K"
    equal_budget = data["budget_units"].groupby(["pool", "candidate_count"], as_index=False).agg(
        audit_time_seconds=("audit_fit_seconds", "mean"),
        normalized_selected_gain=("homogeneous_normalized_selected_gain", "mean"),
        endpoint_count=("task", "nunique"),
        mean_eligible_candidates=("eligible_candidate_count", "mean"),
    )
    equal_budget["design"] = "equal_budget"
    panel_d = pd.concat([equal_k, equal_budget], ignore_index=True)
    tables = {"Panel A": panel_a, "Panel B": panel_b, "Panel C": panel_c, "Panel D": panel_d}
    for name, table in tables.items():
        table.to_csv(SOURCE / f"Figure_7_{name.replace(' ', '_')}_source.csv", index=False)
    path = SOURCE / "Figure_7_source_data.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, table in tables.items():
            table.to_excel(writer, sheet_name=name, index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F5597")
            for column in range(1, sheet.max_column + 1):
                values = [sheet.cell(row, column).value for row in range(1, min(sheet.max_row, 250) + 1)]
                width = min(max(len(str(value)) for value in values if value is not None) + 2, 34)
                sheet.column_dimensions[get_column_letter(column)].width = max(10, width)


def mean_by_endpoint(summary: pd.DataFrame, metric: str) -> pd.DataFrame:
    source = summary.loc[
        summary.design.eq("equal_K") & summary.anchor_scheme.eq("shared_morgan_linear")
    ]
    return source.groupby(["pool", "task", "task_type", "candidate_count"], as_index=False)[f"{metric}_mean"].mean()


def figure7(data: dict[str, pd.DataFrame]) -> None:
    summary = data["summary"]
    stability = data["stability"]
    units = data["units"]
    budget_units = data["budget_units"]
    fig = plt.figure(figsize=(16, 10), constrained_layout=True)
    gs = fig.add_gridspec(2, 2, width_ratios=[1.04, 1.0], height_ratios=[1.0, 1.08])

    # A: endpoint-level opportunity and realised gain at K=32.
    ax = fig.add_subplot(gs[0, 0])
    source = summary.loc[
        summary.design.eq("equal_K") & summary.anchor_scheme.eq("shared_morgan_linear")
        & summary.candidate_count.eq(32)
    ].copy()
    rows = []
    for task in TASKS:
        for pool in POOLS:
            row = source.loc[source.task.eq(task) & source.pool.eq(pool)].iloc[0]
            rows.append((task, pool, row.homogeneous_normalized_selected_gain_mean,
                         row.homogeneous_normalized_oracle_opportunity_mean))
    y = np.arange(len(rows))[::-1]
    for pos, (task, pool, selected, opportunity) in zip(y, rows):
        color = COLORS[pool]
        ax.plot([selected, opportunity], [pos, pos], color=color, linewidth=1.2, alpha=0.8)
        ax.scatter(opportunity, pos, s=28, facecolors="white", edgecolors=color, linewidth=1.1, zorder=3)
        ax.scatter(selected, pos, s=26, color=color, marker="s", zorder=3)
    labels = [f"{TASK_LABEL[t]} · {POOL_SHORT[p]}" for t, p, _, _ in rows][::-1]
    ax.set_yticks(y, labels[::-1])
    ax.axhline(len(rows) - 9.5, color="#8B95A5", linewidth=0.8)
    ax.text(0.99, 0.98, "Classification", transform=ax.transAxes, ha="right", va="top", color="#596273")
    ax.text(0.99, 0.47, "Regression", transform=ax.transAxes, ha="right", va="top", color="#596273")
    ax.axvline(0, color="#687386", linewidth=0.8)
    ax.set_xlabel("Paired homogeneous-normalized gain")
    ax.set_title("Endpoint-level opportunity and realised gain", loc="left", fontweight="bold")
    clean(ax, "x")
    ax.legend(handles=[
        Line2D([0], [0], marker="o", markerfacecolor="white", markeredgecolor="#333", color="none", label="Audit-best opportunity"),
        Line2D([0], [0], marker="s", color="#333", linestyle="none", label="Validation-selected gain"),
    ], frameon=False, ncol=2, loc="lower right")
    panel_label(ax, "A")

    # B: composition-by-K ladder, with separate classification/regression x bands.
    ax = fig.add_subplot(gs[0, 1])
    class_x = np.arange(4)
    reg_x = np.arange(4) + 5
    for pool in POOLS:
        p = source.iloc[0:0]
        full = summary.loc[
            summary.design.eq("equal_K") & summary.anchor_scheme.eq("shared_morgan_linear")
            & summary.pool.eq(pool)
        ]
        for task_type, x in [("classification", class_x), ("regression", reg_x)]:
            grouped = full.loc[full.task_type.eq(task_type)].groupby("candidate_count", as_index=False).agg(
                selected=("homogeneous_normalized_selected_gain_mean", "mean"),
                gap=("homogeneous_normalized_cross_fitted_gap_mean", "mean"),
            ).set_index("candidate_count").reindex(KS)
            ax.plot(x, grouped.selected, color=COLORS[pool], marker="o", markersize=4, linewidth=1.5)
            ax.plot(x, grouped.gap, color=COLORS[pool], marker="o", markerfacecolor="white",
                    markersize=4, linewidth=1.1, linestyle="--")
    ax.axvline(4, color="#8B95A5", linewidth=0.8)
    ax.axhline(0, color="#687386", linewidth=0.8)
    ax.set_xticks(np.r_[class_x, reg_x], [str(k) for k in KS] * 2)
    header_box = dict(facecolor="white", edgecolor="none", alpha=0.86, pad=1.5)
    ax.text(1.5, 0.985, "Classification", transform=ax.get_xaxis_transform(), ha="center", va="top",
            fontweight="bold", bbox=header_box)
    ax.text(6.5, 0.985, "Regression", transform=ax.get_xaxis_transform(), ha="center", va="top",
            fontweight="bold", bbox=header_box)
    ax.set_xlabel("Candidate-pool size, K")
    ax.set_ylabel("Paired homogeneous-normalized value")
    ax.set_title("Composition-by-K ladder", loc="left", fontweight="bold")
    clean(ax, "y")
    handles = [Line2D([0], [0], color=COLORS[p], lw=2, label=POOL_SHORT[p]) for p in POOLS]
    handles += [
        Line2D([0], [0], color="#333", marker="o", lw=1.5, label="Selected gain"),
        Line2D([0], [0], color="#333", marker="o", markerfacecolor="white", ls="--", label="Cross-fitted gap"),
    ]
    ax.legend(handles=handles, frameon=False, ncol=2, loc="best")
    panel_label(ax, "B")

    # C: CAHit@3 heatmap and entropy strip.
    ax = fig.add_subplot(gs[1, 0])
    primary = summary.loc[
        summary.design.eq("equal_K") & summary.anchor_scheme.eq("shared_morgan_linear")
    ]
    row_index = pd.MultiIndex.from_product([TASKS, POOLS], names=["task", "pool"])
    heat = primary.pivot_table(index=["task", "pool"], columns="candidate_count", values="chance_adjusted_hit3_mean").reindex(row_index)[KS]
    entropy = stability.set_index(["task", "pool", "candidate_count"])["candidate_selection_entropy_normalized"]
    entropy32 = pd.Series([entropy.get((task, pool, 32), np.nan) for task, pool in row_index], index=row_index)
    matrix = np.c_[heat.to_numpy(float), entropy32.to_numpy(float)]
    image = vector_heatmap(ax, matrix, cmap="RdYlBu", vmin=-1, vmax=1)
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            value = matrix[i, j]
            if np.isfinite(value):
                ax.text(j + 0.5, i + 0.5, f"{value:.2f}", ha="center", va="center", fontsize=6,
                        color="white" if abs(value) > 0.58 else "#273142")
    ax.set_xticks(np.arange(5) + 0.5, ["K=4", "K=8", "K=16", "K=32", "Entropy\n(K=32)"])
    ax.set_yticks(np.arange(len(row_index)) + 0.5, [f"{TASK_LABEL[t]} · {POOL_SHORT[p]}" for t, p in row_index])
    ax.axvline(4, color="white", linewidth=2)
    ax.set_title("Ranking fidelity and selection stability", loc="left", fontweight="bold")
    cbar = fig.colorbar(image, ax=ax, fraction=0.03, pad=0.02)
    if cbar.solids is not None:
        cbar.solids.set_rasterized(False)
    cbar.set_label("CAHit@3 / normalized entropy")
    panel_label(ax, "C")

    # D: downstream budget-benefit frontier.
    ax = fig.add_subplot(gs[1, 1])
    equal_k = units.loc[
        units.design.eq("equal_K") & units.anchor_scheme.eq("shared_morgan_linear")
    ].groupby(["pool", "candidate_count"], as_index=False).agg(
        time=("audit_fit_seconds", "mean"),
        gain=("homogeneous_normalized_selected_gain", "mean"),
    )
    equal_budget = budget_units.groupby(["pool", "candidate_count"], as_index=False).agg(
        time=("audit_fit_seconds", "mean"), gain=("homogeneous_normalized_selected_gain", "mean")
    )
    equal_k["design"] = "Equal K"
    equal_budget["design"] = "Equal budget"
    points = pd.concat([equal_k, equal_budget], ignore_index=True)
    for pool in POOLS:
        for design, marker, linestyle in [("Equal K", "o", "-"), ("Equal budget", "D", "--")]:
            group = points.loc[points.pool.eq(pool) & points.design.eq(design)].sort_values("candidate_count")
            ax.plot(group.time, group.gain, color=COLORS[pool], marker=marker, linestyle=linestyle,
                    linewidth=1.3, markersize=4)
            for _, row in group.iterrows():
                label = f"{int(row.candidate_count)}" if design == "Equal K" else f"B{int(row.candidate_count)}"
                offset = (3, 3) if design == "Equal K" else (3, -8)
                ax.annotate(label, (row.time, row.gain), xytext=offset,
                            textcoords="offset points", fontsize=6, color=COLORS[pool])
    finite = points.dropna(subset=["time", "gain"]).loc[lambda x: x.gain.gt(0)].sort_values("time")
    frontier = []
    best = -np.inf
    for _, row in finite.iterrows():
        if row.gain > best:
            frontier.append(row)
            best = row.gain
    if frontier:
        f = pd.DataFrame(frontier)
        ax.step(f.time, f.gain, where="post", color="#202632", linewidth=1.0, alpha=0.75, label="Pareto frontier")
    ax.set_xscale("log")
    ax.axhline(0, color="#687386", linewidth=0.8)
    ax.set_xlabel("Downstream audit time per outer unit (s, log scale)")
    ax.set_ylabel("Paired homogeneous-normalized selected gain")
    ax.set_title("Downstream budget–benefit frontier", loc="left", fontweight="bold")
    clean(ax, "both")
    handles = [Line2D([0], [0], color=COLORS[p], lw=2, label=POOL_SHORT[p]) for p in POOLS]
    handles += [
        Line2D([0], [0], color="#333", marker="o", label="Equal K"),
        Line2D([0], [0], color="#333", marker="D", ls="--", label="Equal budget"),
        Line2D([0], [0], color="#202632", drawstyle="steps-post", label="Pareto frontier"),
    ]
    ax.legend(handles=handles, frameon=False, ncol=2, loc="best")
    ax.text(0.02, 0.02, "Point labels: K; B = equal budget", transform=ax.transAxes,
            fontsize=6.5, color="#596273", ha="left", va="bottom")
    panel_label(ax, "D")

    fig.suptitle("Candidate-pool composition governs opportunity, selection pressure and downstream efficiency",
                 fontsize=13, fontweight="bold")
    save(fig, OUT, "Figure_7_expanded_equal_size_intervention")


def supplementary_s16(data: dict[str, pd.DataFrame]) -> None:
    summary = data["summary"]
    source = summary.loc[
        summary.design.eq("equal_K") & summary.anchor_scheme.eq("shared_morgan_linear")
    ]
    fig, axes = plt.subplots(2, 3, figsize=(14, 8), constrained_layout=True, sharex=True)
    for ax, task in zip(axes.flat, TASKS):
        for pool in POOLS:
            g = source.loc[source.task.eq(task) & source.pool.eq(pool)].set_index("candidate_count").reindex(KS)
            ax.plot(KS, g.homogeneous_normalized_oracle_opportunity_mean, color=COLORS[pool], marker="o", lw=1.4)
            ax.plot(KS, g.homogeneous_normalized_selected_gain_mean, color=COLORS[pool], marker="s", ls="--", lw=1.1)
        ax.axhline(0, color="#687386", lw=0.7)
        ax.set_title(TASK_LABEL[task], fontweight="bold")
        ax.set_xticks(KS)
        clean(ax)
    axes[1, 0].set_xlabel("K")
    axes[1, 1].set_xlabel("K")
    axes[1, 2].set_xlabel("K")
    axes[0, 0].set_ylabel("Normalized gain")
    axes[1, 0].set_ylabel("Normalized gain")
    handles = [Line2D([0], [0], color=COLORS[p], label=POOL_SHORT[p]) for p in POOLS]
    handles += [Line2D([0], [0], color="#333", marker="o", label="Audit-best"),
                Line2D([0], [0], color="#333", marker="s", ls="--", label="Selected")]
    fig.legend(handles=handles, frameon=False, ncol=5, loc="upper center")
    fig.suptitle("All endpoint × pool × K intervention results", fontsize=12, fontweight="bold")
    save(fig, SUPP, "Supplementary_Figure_S16_all_endpoint_pool_K")


def supplementary_s17(data: dict[str, pd.DataFrame]) -> None:
    source = data["ablation"]
    diversity = data["ablation_diversity"]
    raw = diversity.loc[diversity.transformation.eq("raw"), [
        "pool", "task", "task_type", "candidate_count", "relative_entropy_rank"
    ]]
    source = source.merge(raw, on=["pool", "task", "task_type", "candidate_count"], validate="one_to_one")
    anchor = data["units"].loc[
        data["units"].design.eq("equal_K") & data["units"].anchor_scheme.eq("shared_morgan_linear"),
        ["task", "seed", "outer_fold", "anchor_utility"],
    ].drop_duplicates(["task", "seed", "outer_fold"])
    endpoint_mad = anchor.groupby("task").anchor_utility.apply(
        lambda x: max(float(np.median(np.abs(x - np.median(x)))), 1e-12)
    )
    source["selected_endpoint_MAD"] = source.apply(
        lambda row: row.selected_model_gain_mean / endpoint_mad[row.task], axis=1
    )
    source["opportunity_endpoint_MAD"] = source.apply(
        lambda row: row.oracle_opportunity_gain_mean / endpoint_mad[row.task], axis=1
    )
    source["gap_endpoint_MAD"] = source.apply(
        lambda row: row.cross_fitted_selection_gap_mean / endpoint_mad[row.task], axis=1
    )
    metrics = [
        ("selected_endpoint_MAD", "Selected gain / endpoint MAD", "diverging"),
        ("opportunity_endpoint_MAD", "Audit-best opportunity / endpoint MAD", "diverging"),
        ("gap_endpoint_MAD", "Cross-fitted gap / endpoint MAD", "diverging"),
        ("chance_adjusted_hit3_mean", "CAHit@3", "bounded"),
        ("chance_adjusted_mrr_mean", "Normalized MRR", "bounded"),
        ("relative_entropy_rank", "Relative entropy rank", "unit"),
    ]
    fig, axes = plt.subplots(2, 3, figsize=(15, 8), constrained_layout=True)
    pools = ["Classical multiview", "+ChemBERTa", "+MoLFormer", "+D-MPNN", "Full modern-augmented"]
    index = pd.MultiIndex.from_product([TASKS, [16, 32]], names=["task", "candidate_count"])
    for ax, (metric, label, scale) in zip(axes.flat, metrics):
        table = source.pivot_table(index=["task", "candidate_count"], columns="pool", values=metric).reindex(index)[pools]
        values = table.to_numpy(float)
        if scale == "diverging":
            vmax = max(float(np.nanquantile(np.abs(values), 0.95)), 1e-9)
            kwargs = {"cmap": "RdBu_r", "vmin": -vmax, "vmax": vmax}
        elif scale == "bounded":
            kwargs = {"cmap": "RdYlBu", "vmin": -1, "vmax": 1}
        else:
            kwargs = {"cmap": "viridis", "vmin": 0, "vmax": 1}
        image = vector_heatmap(ax, values, **kwargs)
        for i in range(values.shape[0]):
            for j in range(values.shape[1]):
                if np.isfinite(values[i, j]):
                    ax.text(j + 0.5, i + 0.5, f"{values[i, j]:.2f}", ha="center", va="center", fontsize=5.1,
                            color="white" if abs(values[i, j]) > 0.65 * max(abs(kwargs["vmin"]), abs(kwargs["vmax"])) else "#273142")
        ax.set_xticks(np.arange(len(pools)) + 0.5, ["Classical", "+ChemBERTa", "+MoLFormer", "+D-MPNN", "Full"], rotation=35, ha="right")
        ax.set_yticks(np.arange(len(index)) + 0.5, [f"{TASK_LABEL[t]} · K={k}" for t, k in index])
        ax.set_title(label, loc="left", fontweight="bold")
        cbar = fig.colorbar(image, ax=ax, fraction=0.035, pad=0.02)
        if cbar.solids is not None:
            cbar.solids.set_rasterized(False)
    fig.suptitle("Fixed-K modern-component ablations", fontsize=12, fontweight="bold")
    save(fig, SUPP, "Supplementary_Figure_S17_modern_component_ablations")


def supplementary_s18(data: dict[str, pd.DataFrame]) -> None:
    table = data["normalization"]
    source = table.loc[table.metric.eq("selected_model_gain") & table.candidate_count.eq(32)].copy()
    rows = [(a, n) for a in source.anchor_scheme.unique() for n in ["raw", "endpoint_MAD", "homogeneous_audit_best"]]
    columns = pd.MultiIndex.from_product([TASKS, POOLS])
    matrix = np.full((len(rows), len(columns)), np.nan)
    for i, (anchor, norm) in enumerate(rows):
        for j, (task, pool) in enumerate(columns):
            hit = source.loc[source.anchor_scheme.eq(anchor) & source.normalization.eq(norm)
                             & source.task.eq(task) & source.pool.eq(pool)]
            if len(hit):
                matrix[i, j] = hit.iloc[0].normalized_mean
    fig, ax = plt.subplots(figsize=(15, 5.8), constrained_layout=True)
    direction = np.sign(matrix)
    image = vector_heatmap(ax, direction, cmap="RdBu_r", vmin=-1, vmax=1)
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            if np.isfinite(matrix[i, j]):
                ax.text(j + 0.5, i + 0.5, f"{matrix[i, j]:.2g}", ha="center", va="center", fontsize=5.2,
                        color="white" if abs(direction[i, j]) > 0 else "#273142")
    ax.set_yticks(np.arange(len(rows)) + 0.5, [f"{a}\n{n}" for a, n in rows])
    ax.set_xticks(np.arange(len(columns)) + 0.5, [f"{TASK_LABEL[t]}\n{POOL_SHORT[p]}" for t, p in columns], rotation=55, ha="right")
    ax.set_title("Anchor and normalization direction sensitivity at K=32", loc="left", fontweight="bold")
    cbar = fig.colorbar(image, ax=ax, fraction=0.025, pad=0.02, ticks=[-1, 0, 1])
    if cbar.solids is not None:
        cbar.solids.set_rasterized(False)
    cbar.ax.set_yticklabels(["negative", "zero", "positive"])
    cbar.set_label("Direction; cell text gives scale-specific mean")
    save(fig, SUPP, "Supplementary_Figure_S18_anchor_normalization_sensitivity")


def supplementary_s19() -> None:
    root = DATA / "composition_split_loop"
    summary = pd.read_csv(root / "Paper31_composition_split_summary.csv")
    diversity = pd.read_csv(root / "Paper31_composition_split_diversity.csv")
    stability = pd.read_csv(root / "Paper31_composition_split_stability.csv")
    metrics = [
        ("homogeneous_normalized_selected_gain_mean", "Normalized selected gain"),
        ("homogeneous_normalized_cross_fitted_gap_mean", "Normalized cross-fitted gap"),
        ("chance_adjusted_hit3_mean", "CAHit@3"),
        ("chance_adjusted_mrr_mean", "Normalized MRR"),
    ]
    fig, axes = plt.subplots(2, 3, figsize=(15, 8), constrained_layout=True)
    for ax, (metric, label) in zip(axes.flat[:4], metrics):
        for pool in POOLS:
            for regime, marker, ls in [("scaffold", "o", "-"), ("similarity_cluster", "s", "--")]:
                g = summary.loc[summary.pool.eq(pool) & summary.split_regime.eq(regime)].groupby("candidate_count")[metric].mean().reindex(KS)
                ax.plot(KS, g, color=COLORS[pool], marker=marker, ls=ls, lw=1.2)
        ax.set_xticks(KS)
        ax.set_xlabel("K")
        ax.set_ylabel(label)
        clean(ax)
    ax = axes.flat[4]
    raw = diversity.loc[diversity.transformation.eq("raw")]
    for pool in POOLS:
        for regime, marker, ls in [("scaffold", "o", "-"), ("similarity_cluster", "s", "--")]:
            g = raw.loc[raw.pool.eq(pool) & raw.split_regime.eq(regime)].groupby("candidate_count").relative_entropy_rank.mean().reindex(KS)
            ax.plot(KS, g, color=COLORS[pool], marker=marker, ls=ls, lw=1.2)
    ax.set_xticks(KS)
    ax.set_xlabel("K")
    ax.set_ylabel("Relative entropy rank")
    clean(ax)
    ax = axes.flat[5]
    for pool in POOLS:
        for regime, marker, ls in [("scaffold", "o", "-"), ("similarity_cluster", "s", "--")]:
            g = stability.loc[stability.pool.eq(pool) & stability.split_regime.eq(regime)].groupby(
                "candidate_count"
            ).candidate_selection_entropy_normalized.mean().reindex(KS)
            ax.plot(KS, g, color=COLORS[pool], marker=marker, ls=ls, lw=1.2)
    ax.set_xticks(KS)
    ax.set_xlabel("K")
    ax.set_ylabel("Normalized selection entropy")
    clean(ax)
    handles = [Line2D([0], [0], color=COLORS[p], label=POOL_SHORT[p]) for p in POOLS]
    handles += [Line2D([0], [0], color="#333", marker="o", label="Scaffold"),
                Line2D([0], [0], color="#333", marker="s", ls="--", label="Tanimoto component")]
    fig.legend(handles=handles, frameon=False, ncol=5, loc="upper center", bbox_to_anchor=(0.5, 1.025))
    fig.suptitle("Composition effects across two evaluated split regimes", fontsize=12, fontweight="bold", y=1.065)
    save(fig, SUPP, "Supplementary_Figure_S19_composition_split_regime")


def supplementary_s20(data: dict[str, pd.DataFrame]) -> None:
    stability = data["stability"]
    row_index = pd.MultiIndex.from_product([TASKS, POOLS], names=["task", "pool"])
    entropy = stability.pivot_table(index=["task", "pool"], columns="candidate_count",
                                    values="candidate_selection_entropy_normalized").reindex(row_index)[KS]
    modal = stability.pivot_table(index=["task", "pool"], columns="candidate_count",
                                  values="modal_candidate_share").reindex(row_index)[KS]
    fig, axes = plt.subplots(1, 2, figsize=(14, 7.5), constrained_layout=True)
    for ax, table, title, cmap in [
        (axes[0], entropy, "Normalized selection entropy", "viridis"),
        (axes[1], modal, "Modal-candidate proportion", "magma_r"),
    ]:
        image = vector_heatmap(ax, table.to_numpy(float), cmap=cmap, vmin=0, vmax=1)
        ax.set_xticks(np.arange(4) + 0.5, [f"K={k}" for k in KS])
        ax.set_yticks(np.arange(len(row_index)) + 0.5, [f"{TASK_LABEL[t]} · {POOL_SHORT[p]}" for t, p in row_index])
        ax.set_title(title, loc="left", fontweight="bold")
        cbar = fig.colorbar(image, ax=ax, fraction=0.03, pad=0.02)
        if cbar.solids is not None:
            cbar.solids.set_rasterized(False)
    save(fig, SUPP, "Supplementary_Figure_S20_selection_frequency_entropy")


def supplementary_s21(data: dict[str, pd.DataFrame]) -> None:
    units = data["units"]
    budget = data["budget_units"]
    equal_k = units.loc[units.design.eq("equal_K") & units.anchor_scheme.eq("shared_morgan_linear")].copy()
    equal_k["comparison"] = "Equal K"
    budget = budget.copy()
    budget["comparison"] = "Equal budget"
    source = pd.concat([equal_k, budget], ignore_index=True)
    fig, axes = plt.subplots(2, 2, figsize=(13, 9), constrained_layout=True)
    for pool in POOLS:
        for design, marker, ls in [("Equal K", "o", "-"), ("Equal budget", "D", "--")]:
            g = source.loc[source.pool.eq(pool) & source.comparison.eq(design)].groupby("candidate_count").agg(
                gain=("homogeneous_normalized_selected_gain", "mean"),
                count=("eligible_candidate_count", "mean"),
                time=("audit_fit_seconds", "mean"),
            ).sort_index()
            axes.flat[0].plot(g.index, g.gain, color=COLORS[pool], marker=marker, ls=ls, lw=1.3)
            axes.flat[1].plot(g.time, g.gain, color=COLORS[pool], marker=marker, ls=ls, lw=1.3)
            cahit = source.loc[source.pool.eq(pool) & source.comparison.eq(design)].groupby(
                "candidate_count"
            ).chance_adjusted_hit3.mean().sort_index()
            axes.flat[2].plot(cahit.index, cahit, color=COLORS[pool], marker=marker, ls=ls, lw=1.3)
    equal_stability = data["stability"].copy()
    budget_stability = data["budget_stability"].copy()
    for pool in POOLS:
        for table, marker, ls in [(equal_stability, "o", "-"), (budget_stability, "D", "--")]:
            g = table.loc[table.pool.eq(pool)].groupby("candidate_count").candidate_selection_entropy_normalized.mean().sort_index()
            axes.flat[3].plot(g.index, g, color=COLORS[pool], marker=marker, ls=ls, lw=1.3)
    axes.flat[0].set_xticks(KS)
    axes.flat[0].set_xlabel("Nominal K")
    axes.flat[0].set_ylabel("Normalized selected gain")
    axes.flat[1].set_xscale("log")
    axes.flat[1].set_xlabel("Downstream audit time (s, log scale)")
    axes.flat[1].set_ylabel("Normalized selected gain")
    axes.flat[2].set_xticks(KS)
    axes.flat[2].set_xlabel("Nominal K")
    axes.flat[2].set_ylabel("CAHit@3")
    axes.flat[3].set_xticks(KS)
    axes.flat[3].set_xlabel("Nominal K")
    axes.flat[3].set_ylabel("Normalized selection entropy")
    for ax in axes.flat:
        clean(ax)
    handles = [Line2D([0], [0], color=COLORS[p], label=POOL_SHORT[p]) for p in POOLS]
    handles += [Line2D([0], [0], color="#333", marker="o", label="Equal K"),
                Line2D([0], [0], color="#333", marker="D", ls="--", label="Equal budget")]
    fig.legend(handles=handles, frameon=False, ncol=5, loc="upper center")
    fig.suptitle("Equal-K versus equal-downstream-budget comparison", fontsize=12, fontweight="bold")
    save(fig, SUPP, "Supplementary_Figure_S21_equalK_equal_budget")


def main() -> None:
    style()
    data = load()
    export_figure7_source(data)
    figure7(data)
    supplementary_s16(data)
    supplementary_s17(data)
    supplementary_s18(data)
    supplementary_s19()
    supplementary_s20(data)
    supplementary_s21(data)
    print(OUT)
    print(SUPP)


if __name__ == "__main__":
    main()
