from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "output" / "paper31_expanded_intervention_20260717"
OUT = DATA / "supplementary_tables"


def compact_registry() -> pd.DataFrame:
    table = pd.read_csv(DATA / "Paper31_candidate_registry.csv")
    columns = ["pool", "pool_order", "candidate", "representation", "family", "source"]
    return table[columns].sort_values(["pool", "pool_order"])


def ablation_table() -> pd.DataFrame:
    table = pd.read_csv(DATA / "Paper31_component_ablation_summary.csv")
    diversity = pd.read_csv(DATA / "Paper31_component_ablation_diversity.csv")
    diversity = diversity.loc[diversity.transformation.eq("raw"), [
        "pool", "task", "task_type", "candidate_count", "entropy_rank",
        "participation_rank", "relative_entropy_rank", "ledoit_wolf_shrinkage_alpha",
    ]]
    table = table.merge(diversity, on=["pool", "task", "task_type", "candidate_count"], validate="one_to_one")
    columns = [
        "pool", "task", "task_type", "candidate_count",
        "oracle_opportunity_gain_mean", "oracle_opportunity_gain_low", "oracle_opportunity_gain_high",
        "selected_model_gain_mean", "selected_model_gain_low", "selected_model_gain_high",
        "cross_fitted_selection_gap_mean", "cross_fitted_selection_gap_low", "cross_fitted_selection_gap_high",
        "chance_adjusted_hit3_mean", "chance_adjusted_mrr_mean", "audit_fit_seconds_mean",
        "entropy_rank", "participation_rank", "relative_entropy_rank", "ledoit_wolf_shrinkage_alpha",
    ]
    return table[columns].sort_values(["task", "candidate_count", "pool"])


def ablation_frequency_table() -> pd.DataFrame:
    return pd.read_csv(DATA / "Paper31_component_ablation_selection_frequency.csv").sort_values(
        ["task", "candidate_count", "pool", "selected_proportion"],
        ascending=[True, True, True, False],
    )


def sensitivity_table() -> pd.DataFrame:
    table = pd.read_csv(DATA / "Paper31_anchor_normalization_sensitivity.csv")
    return table.sort_values([
        "task", "candidate_count", "pool", "anchor_scheme", "normalization", "metric"
    ])


def sensitivity_direction_table() -> pd.DataFrame:
    return pd.read_csv(DATA / "Paper31_anchor_normalization_direction_concordance.csv").sort_values(
        ["anchor_scheme", "normalization"]
    )


def split_table() -> pd.DataFrame:
    root = DATA / "composition_split_loop"
    summary = pd.read_csv(root / "Paper31_composition_split_summary.csv")
    direction = pd.read_csv(root / "Paper31_composition_split_direction_concordance.csv")
    summary["record_type"] = "regime_summary"
    direction["record_type"] = "direction_concordance"
    summary.to_csv(OUT / "Table_S35a_split_regime_summary.csv", index=False)
    direction.to_csv(OUT / "Table_S35b_direction_concordance.csv", index=False)
    return summary


def split_concordance_table() -> pd.DataFrame:
    return pd.read_csv(
        DATA / "composition_split_loop" / "Paper31_composition_split_direction_concordance.csv"
    ).sort_values(["task", "pool", "metric", "contrast"])


def split_stability_table() -> pd.DataFrame:
    return pd.read_csv(
        DATA / "composition_split_loop" / "Paper31_composition_split_stability.csv"
    ).sort_values(["split_regime", "task", "pool", "candidate_count"])


def stability_runtime_table() -> pd.DataFrame:
    stability = pd.read_csv(DATA / "Paper31_selection_stability.csv")
    summary = pd.read_csv(DATA / "Paper31_endpoint_pool_K_summary.csv")
    runtime = summary.loc[
        summary.design.eq("equal_K") & summary.anchor_scheme.eq("shared_morgan_linear"),
        [
            "pool", "task", "task_type", "candidate_count", "audit_fit_seconds_mean",
            "audit_fit_seconds_low", "audit_fit_seconds_high", "selected_gain_per_audit_hour_mean",
        ],
    ]
    return stability.merge(
        runtime, on=["pool", "task", "task_type", "candidate_count"], validate="one_to_one"
    ).sort_values(["task", "pool", "candidate_count"])


def selection_frequency_table() -> pd.DataFrame:
    return pd.read_csv(DATA / "Paper31_selection_frequencies.csv").sort_values(
        ["task", "pool", "candidate_count", "frequency_level", "selected_proportion"],
        ascending=[True, True, True, True, False],
    )


def budget_stability_table() -> pd.DataFrame:
    return pd.read_csv(DATA / "Paper31_equal_budget_selection_stability.csv").sort_values(
        ["task", "pool", "candidate_count"]
    )


def budget_diversity_table() -> pd.DataFrame:
    return pd.read_csv(DATA / "Paper31_equal_budget_effective_diversity.csv").sort_values(
        ["task", "pool", "candidate_count", "transformation"]
    )


def write_sheet(book: Workbook, name: str, frame: pd.DataFrame, note: str) -> None:
    sheet = book.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    sheet.cell(1, 1, note)
    sheet.cell(1, 1).font = Font(name="Arial", size=10, italic=True, color="44546A")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(frame.columns)))
    for col, value in enumerate(frame.columns, start=1):
        cell = sheet.cell(3, col, value)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=4):
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, col_index, None if pd.isna(value) else value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3F6FA")
            if isinstance(value, float):
                cell.number_format = "0.0000"
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(frame.columns))}{len(frame) + 3}"
    for index, column in enumerate(frame.columns, start=1):
        sample = [len(str(column))] + [len(str(v)) for v in frame[column].dropna().head(250)]
        width = min(max(sample) + 2, 34)
        sheet.column_dimensions[get_column_letter(index)].width = max(10, width)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    tables = {
        "S32 Registry": (
            compact_registry(),
            "Table S32. Complete frozen expanded intervention registry; order was fixed before expanded outcomes were read.",
        ),
        "S33 Ablations": (
            ablation_table(),
            "Table S33. Fixed-K modern-component ablations; five seed blocks, with outer folds averaged within seed.",
        ),
        "S33 Frequencies": (
            ablation_frequency_table(),
            "Table S33b. Selected representation and learner-family frequencies for fixed-K component ablations.",
        ),
        "S34 Sensitivity": (
            sensitivity_table(),
            "Table S34. Predefined anchor and normalization sensitivity; denominators <= 1e-12 are reported as invalid.",
        ),
        "S34 Direction": (
            sensitivity_direction_table(),
            "Table S34b. Direction concordance relative to the shared Morgan linear anchor on the raw endpoint scale.",
        ),
        "S35 Split loop": (
            split_table(),
            "Table S35. Candidate-composition effects under seeded scaffold and Tanimoto-component splits; this is not external validation.",
        ),
        "S35 Concordance": (
            split_concordance_table(),
            "Table S35b. Direction concordance of prespecified composition-by-K contrasts across the two evaluated split regimes.",
        ),
        "S35 Stability": (
            split_stability_table(),
            "Table S35c. Selection entropy, modal proportion, seed agreement and fold switches under both split regimes.",
        ),
        "S36 Stability runtime": (
            stability_runtime_table(),
            "Table S36. Candidate/representation/family selection stability and measured downstream audit time.",
        ),
        "S36 Frequencies": (
            selection_frequency_table(),
            "Table S36b. Full candidate, representation and learner-family selection frequencies.",
        ),
        "S36 Budget stability": (
            budget_stability_table(),
            "Table S36c. Equal-downstream-budget selection stability; entropy denominator is the largest eligible prefix in the cell.",
        ),
        "S36 Budget diversity": (
            budget_diversity_table(),
            "Table S36d. Equal-budget effective rank uses the candidate prefix common to all 15 outer units in the endpoint/pool/K cell.",
        ),
    }
    csv_names = {
        "S32 Registry": "Table_S32_expanded_intervention_registry.csv",
        "S33 Ablations": "Table_S33_modern_component_ablations.csv",
        "S33 Frequencies": "Table_S33b_component_selection_frequencies.csv",
        "S34 Sensitivity": "Table_S34_anchor_normalization_sensitivity.csv",
        "S34 Direction": "Table_S34b_direction_concordance.csv",
        "S35 Split loop": "Table_S35_split_regime_interaction.csv",
        "S35 Concordance": "Table_S35b_direction_concordance.csv",
        "S35 Stability": "Table_S35c_split_regime_selection_stability.csv",
        "S36 Stability runtime": "Table_S36_selection_stability_runtime.csv",
        "S36 Frequencies": "Table_S36b_selection_frequencies.csv",
        "S36 Budget stability": "Table_S36c_equal_budget_stability.csv",
        "S36 Budget diversity": "Table_S36d_equal_budget_diversity.csv",
    }
    book = Workbook()
    book.remove(book.active)
    for name, (frame, note) in tables.items():
        frame.to_csv(OUT / csv_names[name], index=False)
        write_sheet(book, name, frame, note)
    book.save(OUT / "Supplementary_Tables_S32-S36.xlsx")
    print(OUT / "Supplementary_Tables_S32-S36.xlsx")


if __name__ == "__main__":
    main()
