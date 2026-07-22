from __future__ import annotations

import copy
import json
import shutil
from pathlib import Path

import fitz
from openpyxl import load_workbook


ROOT = Path(r"D:\fzyc\output\paper33_final_minor_revision_20260718")
SUPP = ROOT / "supplementary"
BASE = Path(r"D:\fzyc\output\paper30_submission_package_20260717\supplementary")
NEW = Path(r"D:\fzyc\output\paper31_submission_package_20260717")

BASE_XLSX = BASE / "Additional_file_2_Machine_readable_Supplementary_Tables_S1-S31.xlsx"
NEW_XLSX = NEW / "supplementary_tables" / "Supplementary_Tables_S32-S36.xlsx"
OUT_XLSX = SUPP / "Additional_file_2_Machine_readable_Supplementary_Tables_S1-S36.xlsx"
BASE_PDF = BASE / "Additional_file_3_Supplementary_Figures_S1-S18.pdf"
OUT_PDF = SUPP / "Additional_file_3_Supplementary_Figures_S1-S21.pdf"


def copy_sheet(source, target) -> None:
    for row in source.iter_rows():
        for cell in row:
            destination = target[cell.coordinate]
            destination.value = cell.value
            if cell.has_style:
                destination._style = copy.copy(cell._style)
            if cell.number_format:
                destination.number_format = cell.number_format
            if cell.font:
                destination.font = copy.copy(cell.font)
            if cell.fill:
                destination.fill = copy.copy(cell.fill)
            if cell.border:
                destination.border = copy.copy(cell.border)
            if cell.alignment:
                destination.alignment = copy.copy(cell.alignment)
            if cell.protection:
                destination.protection = copy.copy(cell.protection)
            if cell.hyperlink:
                destination._hyperlink = copy.copy(cell.hyperlink)
            if cell.comment:
                destination.comment = copy.copy(cell.comment)
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
        target.column_dimensions[key].outlineLevel = dimension.outlineLevel
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key].height = dimension.height
        target.row_dimensions[key].hidden = dimension.hidden
        target.row_dimensions[key].outlineLevel = dimension.outlineLevel
    for merged in source.merged_cells.ranges:
        target.merge_cells(str(merged))
    target.freeze_panes = source.freeze_panes
    if source.auto_filter and source.auto_filter.ref:
        target.auto_filter.ref = source.auto_filter.ref
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.sheet_properties = copy.copy(source.sheet_properties)
    target.page_margins = copy.copy(source.page_margins)
    target.page_setup = copy.copy(source.page_setup)
    target.print_options = copy.copy(source.print_options)


def build_xlsx() -> dict:
    shutil.copy2(BASE_XLSX, OUT_XLSX)
    workbook = load_workbook(OUT_XLSX)
    additions = load_workbook(NEW_XLSX, data_only=False)
    for name in additions.sheetnames:
        if name in workbook.sheetnames:
            raise RuntimeError(f"Duplicate sheet: {name}")
        copy_sheet(additions[name], workbook.create_sheet(name))
    workbook.save(OUT_XLSX)
    workbook.close()
    additions.close()

    check = load_workbook(OUT_XLSX, read_only=True, data_only=False)
    formula_errors = []
    formulas = 0
    for sheet in check.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    if any(token in cell.value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
    sheetnames = check.sheetnames
    check.close()
    return {"sheet_count": len(sheetnames), "sheetnames": sheetnames, "formula_count": formulas, "formula_errors": formula_errors}


def build_pdf() -> dict:
    output = fitz.open()
    base = fitz.open(BASE_PDF)
    if base.page_count < 15:
        raise RuntimeError(f"Expected at least 15 base pages, found {base.page_count}")
    output.insert_pdf(base, from_page=0, to_page=14)
    base.close()
    copied = []
    for number in range(16, 22):
        path = next((NEW / "supplementary_figures").glob(f"Supplementary_Figure_S{number}_*.pdf"))
        source = fitz.open(path)
        output.insert_pdf(source)
        source.close()
        copied.append(path.name)
        shutil.copy2(path, SUPP / path.name)
        for extension in ("svg", "png"):
            matches = list((NEW / "supplementary_figures").glob(f"Supplementary_Figure_S{number}_*.{extension}"))
            for match in matches:
                shutil.copy2(match, SUPP / match.name)
    output.save(OUT_PDF, garbage=4, deflate=True)
    output.close()
    check = fitz.open(OUT_PDF)
    pages = check.page_count
    raster_only_pages = []
    for index, page in enumerate(check):
        if not page.get_drawings() and not page.get_text().strip() and page.get_images(full=True):
            raster_only_pages.append(index + 1)
    check.close()
    return {"page_count": pages, "new_pages": copied, "raster_only_pages": raster_only_pages}


def main() -> None:
    SUPP.mkdir(parents=True, exist_ok=True)
    xlsx = build_xlsx()
    pdf = build_pdf()
    report = {"tables": xlsx, "figures": pdf}
    (ROOT / "Supplementary_numbering_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
